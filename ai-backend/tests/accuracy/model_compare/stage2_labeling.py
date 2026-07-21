"""Step 2 Stage 2 — human clause-type labeling sheet (ground-truth collection).

Stage 1 measured AGREEMENT-WITH-CLAUDE (no human truth). Stage 2 collects the
TRUE clause type from a human so correctness can finally be measured. This module
holds the pure, testable logic:
  - stratified_sample: draw a per-type-balanced Arabic subset (every category
    represented, including rare ones — NOT 150 `general`).
  - build_label_rows: shape the visible sheet rows (Claude's label shown =
    Option 2; the model's prediction is NEVER shown, to avoid biasing the human).
  - validate_human_labels: accept only labels in the canonical 17-set.
  - write_labeling_xlsx / read_filled_xlsx: XLSX I/O with a dropdown + a hidden
    row_id->clause_id map (openpyxl imported lazily, so this module loads without
    the eval deps).
Eval-only; no model calls, no billing.
"""
from __future__ import annotations

import collections
import hashlib
import random
from typing import Any

from tests.accuracy.model_compare.classification_scorer import LABELS_17

# Visible columns the human sees/fills (the model's prediction is intentionally absent).
VISIBLE_COLUMNS = ["row_id", "contract", "language", "clause_text",
                   "claude_label", "human_label", "notes"]


def stratified_allocation(counts: dict[str, int], n_total: int, floor: int = 6) -> dict[str, int]:
    """How many clauses to draw per class: each class first gets `min(support, floor)`
    (so rare classes are fully/represented), then the remainder is distributed
    proportionally to each class's LEFTOVER capacity (largest-remainder), capped at
    support, summing to EXACTLY min(n_total, total_support)."""
    labels = [l for l in counts if counts[l] > 0]
    n_total = min(n_total, sum(counts[l] for l in labels))
    alloc = {l: min(counts[l], floor) for l in labels}
    remaining = n_total - sum(alloc.values())

    if remaining < 0:  # floors already exceed n_total -> trim largest allocations (keep >=1)
        order = sorted(labels, key=lambda l: (-alloc[l], l))
        i = 0
        while remaining < 0 and any(alloc[l] > 1 for l in labels):
            l = order[i % len(order)]
            if alloc[l] > 1:
                alloc[l] -= 1
                remaining += 1
            i += 1
        return alloc

    cap = {l: counts[l] - alloc[l] for l in labels}
    weight_total = sum(cap.values())
    if remaining == 0 or weight_total == 0:
        return alloc
    quota = {l: remaining * cap[l] / weight_total for l in labels}
    for l in labels:
        add = min(cap[l], int(quota[l]))
        alloc[l] += add
        cap[l] -= add
    short = n_total - sum(alloc.values())
    frac_order = sorted((l for l in labels if cap[l] > 0),
                        key=lambda l: (-(quota[l] - int(quota[l])), l))
    i = 0
    while short > 0 and any(cap[l] > 0 for l in labels):
        l = frac_order[i % len(frac_order)]
        if cap[l] > 0:
            alloc[l] += 1
            cap[l] -= 1
            short -= 1
        i += 1
    return alloc


def _seed_for(seed: int, label: str) -> int:
    """Deterministic per-class seed (hashlib, NOT builtin hash which is salted)."""
    return (seed * 1_000_003 + int(hashlib.md5(label.encode("utf-8")).hexdigest()[:8], 16)) & 0x7FFFFFFF


def stratified_sample(records: list[dict], n_total: int = 150,
                      floor: int = 6, seed: int = 0) -> tuple[list[dict], dict[str, int]]:
    """Deterministically draw ~n_total records stratified by `y_true`. Within a class,
    a per-class seeded shuffle selects which clauses (reproducible run-to-run).
    Returns (selected_records, allocation_dict)."""
    by_class: dict[str, list[dict]] = collections.defaultdict(list)
    for r in records:
        by_class[str(r["y_true"])].append(r)
    counts = {l: len(v) for l, v in by_class.items()}
    alloc = stratified_allocation(counts, n_total, floor)
    selected: list[dict] = []
    for lab in sorted(by_class):                          # deterministic class order
        pool = sorted(by_class[lab], key=lambda r: str(r["clause_id"]))
        k = min(alloc.get(lab, 0), len(pool))
        selected.extend(random.Random(_seed_for(seed, lab)).sample(pool, k))
    return selected, alloc


def build_label_rows(en_records: list[dict], ar_records: list[dict]) -> list[dict]:
    """Shape the visible labeling rows. `claude_label` is shown (Option 2);
    `human_label`/`notes` start EMPTY; `clause_id` is carried for the hidden map
    only (never a visible column)."""
    rows: list[dict] = []
    for i, r in enumerate(list(en_records) + list(ar_records), start=1):
        rows.append({
            "row_id": f"R{i:04d}",
            "contract": r.get("contract"),
            "language": r["lang"],
            "clause_text": r.get("text", "") or "",
            "claude_label": r.get("y_true"),   # the yardstick, shown to the human
            "human_label": "",                 # <- user fills this
            "notes": "",
            "clause_id": str(r["clause_id"]),   # hidden mapping, not a visible column
        })
    return rows


def validate_human_labels(filled_rows: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
    """Partition filled rows into (labeled, unlabeled, invalid). A human_label is
    valid iff (case-insensitively) it is one of the 17. Returns normalized labeled
    rows (human_label lowercased)."""
    valid = set(LABELS_17)
    labeled, unlabeled, invalid = [], [], []
    for r in filled_rows:
        hl = str(r.get("human_label") or "").strip().lower()
        if not hl:
            unlabeled.append(r)
        elif hl in valid:
            labeled.append({**r, "human_label": hl})
        else:
            invalid.append(r)
    return labeled, unlabeled, invalid


# ------------------------------------------------------------------ XLSX I/O (lazy)
def write_labeling_xlsx(rows: list[dict], path: str, *, labels: list[str] = LABELS_17) -> None:
    """Write the labeling workbook: a `labeling` sheet (with a human_label dropdown
    from the 17), a `Labels` reference sheet, and a HIDDEN `_map` sheet
    (row_id -> clause_id) so we can rejoin after the human fills it. Lazy openpyxl."""
    from openpyxl import Workbook  # noqa: PLC0415 — eval-only
    from openpyxl.styles import Alignment, Font  # noqa: PLC0415
    from openpyxl.worksheet.datavalidation import DataValidation  # noqa: PLC0415

    wb = Workbook()
    ws = wb.active
    ws.title = "labeling"
    ws.append(VISIBLE_COLUMNS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r["row_id"], r["contract"], r["language"], r["clause_text"],
                   r["claude_label"], r["human_label"], r["notes"]])
    # human_label (col F) dropdown limited to the 17. NOTE openpyxl quirk:
    # showDropDown=False actually SHOWS the in-cell arrow.
    dv = DataValidation(type="list", formula1='"' + ",".join(labels) + '"',
                        allow_blank=True, showDropDown=False)
    dv.errorTitle, dv.error = "Invalid label", "Pick one of the 17 clause types (see the Labels tab)."
    dv.promptTitle, dv.prompt = "Clause type", "Choose the TRUE type for this clause."
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(rows) + 1}")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 90
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 30
    for i, r in enumerate(rows, start=2):
        cell = ws.cell(row=i, column=4)  # clause_text
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   readingOrder=2 if r["language"] == "ar" else 0)

    ref = wb.create_sheet("Labels")
    ref.append(["Valid clause_type labels — put ONE of these in human_label:"])
    ref["A1"].font = Font(bold=True)
    for lab in labels:
        ref.append([lab])
    ref.column_dimensions["A"].width = 40

    mp = wb.create_sheet("_map")
    mp.append(["row_id", "clause_id", "language"])
    for r in rows:
        mp.append([r["row_id"], r["clause_id"], r["language"]])
    mp.sheet_state = "hidden"

    wb.save(path)


def read_filled_xlsx(path: str) -> list[dict]:
    """Read a filled labeling workbook back into row dicts, rejoining `clause_id`
    from the hidden `_map` sheet. Lazy openpyxl."""
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, data_only=True)
    ws = wb["labeling"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[idx["row_id"]] is None:
            continue
        rows.append({
            "row_id": raw[idx["row_id"]],
            "contract": raw[idx.get("contract", -1)] if "contract" in idx else None,
            "language": raw[idx["language"]],
            "claude_label": raw[idx["claude_label"]],
            "human_label": raw[idx["human_label"]] or "",
            "notes": raw[idx["notes"]] if "notes" in idx else None,
        })
    cid = {}
    if "_map" in wb.sheetnames:
        m = wb["_map"]
        mh = [c.value for c in m[1]]
        mi = {h: i for i, h in enumerate(mh)}
        for raw in m.iter_rows(min_row=2, values_only=True):
            cid[raw[mi["row_id"]]] = raw[mi["clause_id"]]
    for r in rows:
        r["clause_id"] = cid.get(r["row_id"])
    return rows
